import frappe
import openpyxl
from frappe.utils import cint, getdate, today
from frappe.utils.background_jobs import enqueue, is_job_enqueued
from io import BytesIO

from hmd_agro.hmd_agro.utils.config import get_config


@frappe.whitelist()
def preview_import(file_url):
    """Parse Excel and return preview: matched animals, date range, row counts"""
    dates, animals_data = parse_excel(file_url)

    if not dates:
        frappe.throw("Aucune date trouvee dans le fichier (ligne 1, colonnes B+).")
    if not animals_data:
        frappe.throw("Aucun animal trouve dans le fichier (colonne A).")

    # Build animal mapping: nom_metier -> animal name
    animal_map, duplicate_noms = build_animal_mapping()

    # Match Excel animals to DB
    matched = []
    unmatched = []
    duplicates = []
    seen_duplicates = set()

    for row in animals_data:
        nom = str(row["nom_metier"]).zfill(4)
        if nom in duplicate_noms:
            if nom not in seen_duplicates:
                duplicates.append(nom)
                seen_duplicates.add(nom)
        elif nom in animal_map:
            matched.append({"nom_metier": nom, "animal": animal_map[nom]})
        else:
            unmatched.append(nom)

    # Build candidate list for each duplicate nom_metier
    duplicate_candidates = {}
    for nom in duplicates:
        candidates = frappe.db.get_all("Animal",
            filters={"nom_metier": nom},
            fields=["name", "identification_tn", "identification_fr",
                    "id_lot", "statut", "categorie"]
        )
        duplicate_candidates[nom] = candidates

    # Build lactation info for matched animals
    lactation_info = {}
    for m in matched:
        lacs = get_animal_lactations(m["animal"])
        lactation_info[m["nom_metier"]] = {
            "animal": m["animal"],
            "lactations": len(lacs),
            "details": [
                {
                    "name": l.name,
                    "numero": l.numero_lactation,
                    "debut": str(l.date_debut),
                    "fin": str(l.date_tarissement) if l.date_tarissement else "en cours",
                    "statut": l.statut
                }
                for l in lacs
            ]
        }

    return {
        "date_range": {
            "start": str(dates[0]),
            "end": str(dates[-1]),
            "days": len(dates)
        },
        "total_animals_excel": len(animals_data),
        "matched": len(matched),
        "unmatched": unmatched,
        "duplicates": duplicates,
        "duplicate_candidates": duplicate_candidates,
        "estimated_traites": len(matched) * len(dates) * 2,
        "lactation_info": lactation_info
    }


def _result_key(file_url):
    return f"import_traites_result::{file_url}"


@frappe.whitelist()
def run_import(file_url, keep_original=1, resolutions=None):
    """Enqueue the import as a background job (a full-year file takes ~90s — too long for a
    synchronous web request, which the proxy times out). The job stashes its result in cache;
    the page polls get_import_result() for it. Polling, not realtime — background-job realtime
    delivery proved unreliable, so the result never surfaced."""
    import json
    keep_original = int(keep_original)
    if isinstance(resolutions, str):
        resolutions = json.loads(resolutions) if resolutions else {}
    resolutions = resolutions or {}

    frappe.cache().delete_value(_result_key(file_url))   # clear any prior result
    # Always enqueue with an RQ-assigned unique id — a fixed job_id could linger in the
    # finished/started registry and make a re-import of the same file silently not run.
    enqueue(
        _process_import, queue="long", timeout=6000,
        file_url=file_url, keep_original=keep_original,
        resolutions=resolutions, user=frappe.session.user,
        now=frappe.conf.developer_mode,
    )
    return {"status": "started", "file_url": file_url}


@frappe.whitelist()
def get_import_result(file_url):
    """Poll endpoint: returns the stored result once the background import finishes, else None."""
    return frappe.cache().get_value(_result_key(file_url))


def _process_import(file_url, keep_original, user, resolutions=None):
    """Background job: parse Excel, create Traite records, stash the result in cache."""
    frappe.set_user(user)
    resolutions = resolutions or {}

    try:
        dates, animals_data = parse_excel(file_url)

        if not dates or not animals_data:
            frappe.cache().set_value(_result_key(file_url),
                {"success": False, "error": "Fichier invalide."}, expires_in_sec=3600)
            return

        animal_map, duplicate_noms = build_animal_mapping()

        # Apply user resolutions for ambiguous nom_metier
        for nom, animal_name in resolutions.items():
            if animal_name:
                animal_map[nom] = animal_name
                duplicate_noms.discard(nom)

        summary = {
            "created": 0,
            "overwritten": 0,
            "skipped_no_animal": 0,
            "skipped_no_lactation": 0,
            "skipped_after_sortie": 0,
            "skipped_duplicate": 0,
            "errors": [],
            "lactations_updated": 0
        }

        affected_lactations = set()

        # Calculate total for progress
        total_cells = len(animals_data) * len(dates)
        processed = 0

        for row in animals_data:
            nom = str(row["nom_metier"]).zfill(4)

            # Skip unmatched or duplicate nom_metier
            if nom in duplicate_noms or nom not in animal_map:
                summary["skipped_no_animal"] += len(dates)
                processed += len(dates)
                frappe.publish_realtime("import_traites_progress",
                    {"current": processed, "total": total_cells},
                    user=user)
                continue

            animal_name = animal_map[nom]
            lactations = get_animal_lactations(animal_name)
            animal_lot = frappe.db.get_value("Animal", animal_name, "id_lot")
            # Mirror Traite.validate_animal_present_on_date — refuse to import
            # traites whose date is after the animal's exit. Without this,
            # Excel files containing stale rows for sold/dead cows silently
            # create post-sortie traites (140 such rows found in the audit
            # before this guard was added).
            animal_sortie = frappe.db.get_value("Animal", animal_name, "date_sortie")

            for i, date in enumerate(dates):
                value = row["values"][i] if i < len(row["values"]) else 0
                try:
                    value = float(value) if value else 0
                except (ValueError, TypeError):
                    summary["errors"].append({
                        "animal": nom,
                        "date": str(date),
                        "reason": f"Valeur non numerique: {value}"
                    })
                    processed += 1
                    continue

                # Skip rows whose date is after the animal's exit (same rule as
                # Traite.validate_animal_present_on_date).
                if animal_sortie and getdate(date) > getdate(animal_sortie):
                    summary["skipped_after_sortie"] += 1
                    if value > 0:
                        summary["errors"].append({
                            "animal": nom,
                            "date": str(date),
                            "reason": f"Animal sorti le {animal_sortie} — "
                                      f"date postérieure ignorée"
                        })
                    processed += 1
                    continue

                # Find which lactation covers this date
                lactation = find_lactation_for_date(lactations, date)

                if not lactation:
                    summary["skipped_no_lactation"] += 1
                    if value > 0:
                        summary["errors"].append({
                            "animal": nom,
                            "date": str(date),
                            "reason": "Pas de lactation pour cette date"
                        })
                    processed += 1
                    continue

                # 50/50 split
                half = round(value / 2, 1)

                max_litres = get_config("traite_max_litres", default=60)
                for session in ["MATIN", "SOIR"]:
                    # Validate quantity (0..max_litres per session — same limit as Traite.validate_quantite)
                    qty = half
                    if qty > max_litres:
                        summary["errors"].append({
                            "animal": nom,
                            "date": str(date),
                            "reason": f"Quantite {session} depasse {max_litres}L ({qty})"
                        })
                        continue

                    # Check duplicate
                    existing = frappe.db.exists("Traite", {
                        "animal": animal_name,
                        "date_traite": str(date),
                        "session": session
                    })
                    if existing:
                        if keep_original:
                            summary["skipped_duplicate"] += 1
                            continue
                        else:
                            # Override: update existing traite. Reset brut to the
                            # imported value too — the import is the new measurement.
                            frappe.db.set_value("Traite", existing,
                                {"quantite_litres": qty,
                                 "quantite_litres_brut": qty,
                                 "lactation": lactation.name})
                            summary["overwritten"] += 1
                            affected_lactations.add(lactation.name)
                            continue

                    doc = frappe.get_doc({
                        "doctype": "Traite",
                        "animal": animal_name,
                        "date_traite": str(date),
                        "session": session,
                        "quantite_litres": qty,
                        "quantite_litres_brut": qty,
                        "lactation": lactation.name,
                        "id_lot": animal_lot
                    })
                    doc.flags.ignore_validate = True
                    doc.flags.ignore_links = True
                    doc.flags.skip_lactation_update = True  # final recalc done once at end of import
                    doc.insert(ignore_permissions=True)
                    summary["created"] += 1
                    affected_lactations.add(lactation.name)

                processed += 1

                # Publish progress every 50 cells
                if processed % 50 == 0:
                    frappe.publish_realtime("import_traites_progress",
                        {"current": processed, "total": total_cells},
                        user=user)

            # Commit per animal
            frappe.db.commit()

        # Recalculate production for all affected lactations
        for lac_name in affected_lactations:
            recalculate_lactation_production(lac_name)
        frappe.db.commit()

        summary["lactations_updated"] = len(affected_lactations)
        frappe.cache().set_value(_result_key(file_url),
            {"success": True, "summary": summary}, expires_in_sec=3600)

    except Exception:
        frappe.db.rollback()
        frappe.log_error("Import traites failed")
        frappe.cache().set_value(_result_key(file_url),
            {"success": False, "error": str(frappe.get_traceback())}, expires_in_sec=3600)


def find_lactation_for_date(lactations, date):
    """Find the lactation that covers the given date"""
    for lac in lactations:
        if not lac.date_debut:
            continue
        debut = getdate(lac.date_debut)
        if debut > date:
            continue
        # EN_COURS has no end date — covers up to today
        if not lac.date_tarissement:
            if date <= getdate(today()):
                return lac
        else:
            if date <= getdate(lac.date_tarissement):
                return lac
    return None


def recalculate_lactation_production(lactation_name):
    """Recalculate production totals for a lactation (same logic as Traite.update_lactation_production)"""
    date_debut = frappe.db.get_value("Lactation", lactation_name, "date_debut")

    total = frappe.db.sql("""
        SELECT SUM(quantite_litres)
        FROM `tabTraite`
        WHERE lactation = %s
    """, lactation_name)[0][0] or 0

    pic_window = cint(get_config("pic_production_jours", default=150))
    pic = frappe.db.sql(f"""
        SELECT MAX(daily_total) FROM (
            SELECT SUM(quantite_litres) as daily_total
            FROM `tabTraite`
            WHERE lactation = %s
              AND DATEDIFF(date_traite, %s) <= {pic_window}
            GROUP BY date_traite
        ) as daily
    """, (lactation_name, date_debut))[0][0] or 0

    updates = {
        "production_totale": total,
        "pic_production": pic
    }

    if date_debut:
        total_305 = frappe.db.sql("""
            SELECT SUM(quantite_litres)
            FROM `tabTraite`
            WHERE lactation = %s
              AND DATEDIFF(date_traite, %s) <= 305
        """, (lactation_name, date_debut))[0][0] or 0
        updates["lactation_305j"] = round(total_305, 2)

        init_window = cint(get_config("production_initiale_jours", default=60))
        prod_init = frappe.db.sql(f"""
            SELECT SUM(quantite_litres)
            FROM `tabTraite`
            WHERE lactation = %s
              AND DATEDIFF(date_traite, %s) <= {init_window}
        """, (lactation_name, date_debut))[0][0] or 0
        updates["production_initiale"] = round(prod_init, 2)

        from frappe.utils import date_diff
        jours = date_diff(today(), date_debut)
        if jours > 0:
            updates["moyenne_production"] = round(total / jours, 2)

    frappe.db.set_value("Lactation", lactation_name, updates)


def parse_excel(file_url):
    """Parse the uploaded Excel file, return (dates, animals_data)"""
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    content = file_doc.get_content()
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    # Row 1: date headers starting at col B (index 2)
    dates = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val and hasattr(val, 'date'):
            dates.append(val.date())
        elif val and isinstance(val, str):
            try:
                dates.append(getdate(val))
            except Exception:
                break
        else:
            break

    # Rows 2+: col A = nom_metier, cols B+ = daily values
    animals_data = []
    for row in range(2, ws.max_row + 1):
        nom_metier = ws.cell(row, 1).value
        if nom_metier is None:
            continue

        values = []
        for col_idx in range(2, 2 + len(dates)):
            val = ws.cell(row, col_idx).value
            values.append(val if val is not None else 0)

        animals_data.append({
            "nom_metier": str(nom_metier).zfill(4) if isinstance(nom_metier, int) else str(nom_metier),
            "values": values
        })

    return dates, animals_data


def build_animal_mapping():
    """Map a milk-file animal number to an Animal by EITHER identification_fr (N°Fr)
    or nom_metier (TN suffix) — both zero-padded to 4 to match the file parsing.
    So a file can be keyed by French OR Tunisian numbers (even mixed in one column).

    A number that resolves to >1 distinct animal (e.g. one cow's FR equals another
    cow's TN suffix) is AMBIGUOUS -> returned in `duplicates` so the import's
    resolution dialog asks the user which cow it is (never silently wrong).
    Returns (map, duplicates set)."""
    from collections import defaultdict
    animals = frappe.db.get_all("Animal",
        fields=["name", "nom_metier", "identification_fr"])

    key_to_animals = defaultdict(set)
    for a in animals:
        for val in (a.nom_metier, a.identification_fr):
            if val:
                key_to_animals[str(val).zfill(4)].add(a.name)

    duplicates = {k for k, names in key_to_animals.items() if len(names) > 1}
    mapping = {k: next(iter(names)) for k, names in key_to_animals.items()
               if len(names) == 1}

    return mapping, duplicates


def get_animal_lactations(animal_name):
    """Get all lactations for an animal, sorted by date_debut"""
    return frappe.db.get_all("Lactation",
        filters={"animal": animal_name},
        fields=["name", "numero_lactation", "date_debut", "date_tarissement", "statut"],
        order_by="date_debut asc"
    )
